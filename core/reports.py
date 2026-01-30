# core/reports.py
from __future__ import annotations
import io
from io import BytesIO
import pandas as pd
from datetime import date, datetime

from core.utils import pt_date_to_dt, fmt_id_str, format_currency_br

def excel_quitacoes_colunas_fixas(df: pd.DataFrame) -> bytes:
    """Gera XLSX no layout fixo (sem Aviso/Situação)."""
    if df is None or df.empty:
        return b""

    cols_pdf = [
        "quitacao_data","hospital","atendimento","paciente","convenio",
        "profissional","grau_participacao","data_procedimento",
        "quitacao_guia_amhptiss","quitacao_valor_amhptiss",
        "quitacao_guia_complemento","quitacao_valor_complemento",
    ]
    base = df.copy()
    for c in cols_pdf:
        if c not in base.columns:
            base[c] = ""

    for col in ["quitacao_guia_amhptiss","quitacao_guia_complemento"]:
        base[col] = base[col].apply(fmt_id_str)

    def _to_dt(s):
        d = pt_date_to_dt(s)
        return pd.to_datetime(d) if d else pd.NaT

    base["quitacao_data_x"]     = base["quitacao_data"].apply(_to_dt)
    base["data_procedimento_x"] = base["data_procedimento"].apply(_to_dt)

    base["quitacao_valor_amhptiss_x"]    = pd.to_numeric(base["quitacao_valor_amhptiss"], errors="coerce")
    base["quitacao_valor_complemento_x"] = pd.to_numeric(base["quitacao_valor_complemento"], errors="coerce")

    out = pd.DataFrame({
        "Quitação":               base["quitacao_data_x"],
        "Hospital":               base["hospital"],
        "Atendimento":            base["atendimento"],
        "Paciente":               base["paciente"],
        "Convênio":               base["convenio"],
        "Profissional":           base["profissional"],
        "Grau":                   base["grau_participacao"],
        "Data Proc.":             base["data_procedimento_x"],
        "Guia AMHPTISS":          base["quitacao_guia_amhptiss"],
        "R$ AMHPTISS":            base["quitacao_valor_amhptiss_x"],
        "Guia Compl.":            base["quitacao_guia_complemento"],
        "R$ Compl.":              base["quitacao_valor_complemento_x"],
    })

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet = "Quitações"
        out.to_excel(writer, sheet_name=sheet, index=False)
        ws = writer.sheets[sheet]

        from openpyxl.styles import Alignment, Font
        date_fmt = "dd/mm/yyyy"
        money_fmt = u'[$R$-pt_BR] #,##0.00'

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_quit = headers.index("Quitação") + 1
        idx_proc = headers.index("Data Proc.") + 1
        idx_v1   = headers.index("R$ AMHPTISS") + 1
        idx_v2   = headers.index("R$ Compl.") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_quit, max_col=idx_quit):
            for cell in row:
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_proc, max_col=idx_proc):
            for cell in row:
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_v1, max_col=idx_v1):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_v2, max_col=idx_v2):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

    return buf.getvalue()
